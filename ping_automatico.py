"""
================================================================================
                    PING AUTOMATICO - VERIFICADOR DE HOSTS
================================================================================

Descrição:
    Verifica conectividade de IPs e domínios do arquivo hosts através de ping.
    Gera planilha Excel com resultados coloridos (verde = sucesso, vermelho = falha).
    Execução paralela para maior velocidade.
    Inclui categoria/identificação de cada IP.

Autor:
    Time de Certificação Residencial

Versão: 1.0.0
Data: Janeiro/2026

================================================================================
"""

__author__ = "Time de Certificação Residencial"
__version__ = "1.0.0"
__description__ = "Ping Automático - Verificador de Hosts"

import re
import subprocess
import platform
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Lock para escrita segura no console
print_lock = threading.Lock()

# Configuração do logging
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_filename = f"ping_log_{timestamp}.txt"

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def parse_hosts_file(filepath):
    """
    Extrai IPs e domínios do arquivo hosts com suas categorias
    Retorna lista de tuplas (ip, dominio, categoria)
    
    """
    entries = []
    ip_pattern = re.compile(r'^(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\s+(.+)$')
    hostname_pattern = re.compile(r'^([a-zA-Z0-9\-\.]+)\s+(.+)$')
    
    current_category = "SEM CATEGORIA"
    
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            original_line = line
            line = line.strip()
            
            # Verifica se é um comentário de categoria (linha que começa com # e tem texto)
            if line.startswith('#'):
                # Extrai o texto do comentário como categoria
                category_text = line.lstrip('#').strip()
                if category_text and not category_text.startswith(('10.', '5.', '22.')):
                    # É uma categoria válida (não é um IP comentado)
                    current_category = category_text.upper()
                continue
            
            # Ignora linhas vazias
            if not line:
                continue
            
            # Remove tabs extras e espaços múltiplos
            line = re.sub(r'\s+', ' ', line)
            
            # Tenta extrair IP e domínios
            match = ip_pattern.match(line)
            if match:
                ip = match.group(1)
                domains = match.group(2).strip().split()
                
                # Valida o IP (não pode ter 00.00.00.00)
                if ip == '00.00.00.00':
                    continue
                    
                for domain in domains:
                    domain = domain.strip()
                    if domain and not domain.startswith('-'):
                        entries.append((ip, domain, current_category))
            else:
                # Tenta extrair hostname que pode resolver para IP
                match = hostname_pattern.match(line)
                if match:
                    host = match.group(1)
                    domains = match.group(2).strip().split()
                    for domain in domains:
                        domain = domain.strip()
                        if domain and not domain.startswith('-'):
                            entries.append((host, domain, current_category))
    
    return entries


def ping_host(host, timeout=2):
    """
    Executa ping no host especificado
    Retorna True se sucesso, False se falha
    """
    try:
        # Detecta o sistema operacional
        param = '-n' if platform.system().lower() == 'windows' else '-c'
        timeout_param = '-w' if platform.system().lower() == 'windows' else '-W'
        timeout_val = str(timeout * 1000) if platform.system().lower() == 'windows' else str(timeout)
        
        # Executa o ping
        command = ['ping', param, '1', timeout_param, timeout_val, host]
        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout + 5
        )
        return result.returncode == 0
    except subprocess.TimeoutExpired:
        return False
    except Exception:
        return False


def ping_with_info(args):
    """
    Executa ping e retorna informações formatadas
    Recebe tupla (index, total, ip, domain, category)
    """
    index, total, ip, domain, category = args
    
    # Tenta primeiro o IP
    ip_success = ping_host(ip)
    
    # Se o IP falhou e o domínio é diferente do IP, tenta o domínio
    domain_success = False
    if domain != ip:
        domain_success = ping_host(domain)
    
    success = ip_success or domain_success
    status = "OK" if success else "FALHOU"
    
    # Log thread-safe
    with print_lock:
        msg = f"[{index:4d}/{total}] [{category:25s}] {ip:20s} | {domain:50s} | {status}"
        logger.info(msg)
    
    return {
        'index': index,
        'ip': ip,
        'domain': domain,
        'category': category,
        'ip_success': ip_success,
        'domain_success': domain_success,
        'success': success
    }


def create_excel_report(results, output_file):
    """
    Cria planilha Excel com os resultados do ping
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado Ping"
    
    # Estilos
  
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    success_font = Font(color="006100", bold=True)
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(color="9C0006", bold=True)
    category_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    category_font = Font(color="7F6000", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Cabeçalhos
    headers = ['#', 'Categoria', 'IP', 'Dominio', 'Ping IP', 'Ping Dominio', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Ordena resultados pelo índice original
    results_sorted = sorted(results, key=lambda x: x['index'])
    
    # Dados
    for row, result in enumerate(results_sorted, 2):
        # Número
        ws.cell(row=row, column=1, value=result['index']).alignment = center_align
        ws.cell(row=row, column=1).border = border
        
        # Categoria
        cat_cell = ws.cell(row=row, column=2, value=result['category'])
        cat_cell.alignment = left_align
        cat_cell.border = border
        cat_cell.fill = category_fill
        cat_cell.font = category_font
        
        # IP
        ws.cell(row=row, column=3, value=result['ip']).border = border
        ws.cell(row=row, column=3).alignment = left_align
        
        # Domínio
        ws.cell(row=row, column=4, value=result['domain']).border = border
        ws.cell(row=row, column=4).alignment = left_align
        
        # Ping IP
        ip_cell = ws.cell(row=row, column=5, value="OK" if result['ip_success'] else "FALHOU")
        ip_cell.alignment = center_align
        ip_cell.border = border
        if result['ip_success']:
            ip_cell.fill = success_fill
            ip_cell.font = success_font
        else:
            ip_cell.fill = fail_fill
            ip_cell.font = fail_font
        
        # Ping Domínio
        domain_cell = ws.cell(row=row, column=6, value="OK" if result['domain_success'] else "FALHOU")
        domain_cell.alignment = center_align
        domain_cell.border = border
        if result['domain_success']:
            domain_cell.fill = success_fill
            domain_cell.font = success_font
        else:
            domain_cell.fill = fail_fill
            domain_cell.font = fail_font
        
        # Status Geral
        status_cell = ws.cell(row=row, column=7, value="ONLINE" if result['success'] else "OFFLINE")
        status_cell.alignment = center_align
        status_cell.border = border
        if result['success']:
            status_cell.fill = success_fill
            status_cell.font = success_font
        else:
            status_cell.fill = fail_fill
            status_cell.font = fail_font
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # Adiciona resumo
    total = len(results)
    online = sum(1 for r in results if r['success'])
    offline = total - online
    
    summary_row = len(results) + 3
    ws.cell(row=summary_row, column=1, value="RESUMO:").font = Font(bold=True, size=14)
    ws.cell(row=summary_row + 1, column=1, value=f"Total de hosts: {total}")
    ws.cell(row=summary_row + 2, column=1, value=f"Online: {online}").font = success_font
    ws.cell(row=summary_row + 2, column=1).fill = success_fill
    ws.cell(row=summary_row + 3, column=1, value=f"Offline: {offline}").font = fail_font
    ws.cell(row=summary_row + 3, column=1).fill = fail_fill
    ws.cell(row=summary_row + 4, column=1, value=f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    # Cria aba de resumo por categoria
    ws_cat = wb.create_sheet(title="Resumo por Categoria")
    
    # Agrupa por categoria
    categories = {}
    for r in results:
        cat = r['category']
        if cat not in categories:
            categories[cat] = {'total': 0, 'online': 0, 'offline': 0}
        categories[cat]['total'] += 1
        if r['success']:
            categories[cat]['online'] += 1
        else:
            categories[cat]['offline'] += 1
    
    # Cabeçalhos da aba de categorias
    cat_headers = ['Categoria', 'Total', 'Online', 'Offline', '% Online']
    for col, header in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Dados das categorias
    row = 2
    for cat, stats in sorted(categories.items()):
        ws_cat.cell(row=row, column=1, value=cat).border = border
        ws_cat.cell(row=row, column=1).fill = category_fill
        ws_cat.cell(row=row, column=1).font = category_font
        
        ws_cat.cell(row=row, column=2, value=stats['total']).border = border
        ws_cat.cell(row=row, column=2).alignment = center_align
        
        online_cell = ws_cat.cell(row=row, column=3, value=stats['online'])
        online_cell.border = border
        online_cell.alignment = center_align
        online_cell.fill = success_fill
        online_cell.font = success_font
        
        offline_cell = ws_cat.cell(row=row, column=4, value=stats['offline'])
        offline_cell.border = border
        offline_cell.alignment = center_align
        if stats['offline'] > 0:
            offline_cell.fill = fail_fill
            offline_cell.font = fail_font
        
        pct = (stats['online'] / stats['total'] * 100) if stats['total'] > 0 else 0
        pct_cell = ws_cat.cell(row=row, column=5, value=f"{pct:.1f}%")
        pct_cell.border = border
        pct_cell.alignment = center_align
        if pct >= 80:
            pct_cell.fill = success_fill
            pct_cell.font = success_font
        elif pct >= 50:
            pct_cell.fill = category_fill
        else:
            pct_cell.fill = fail_fill
            pct_cell.font = fail_font
        
        row += 1
    
    # Ajusta largura das colunas da aba de categorias
    ws_cat.column_dimensions['A'].width = 35
    ws_cat.column_dimensions['B'].width = 10
    ws_cat.column_dimensions['C'].width = 10
    ws_cat.column_dimensions['D'].width = 10
    ws_cat.column_dimensions['E'].width = 12
    
    # Salva arquivo
    wb.save(output_file)
    logger.info(f"Planilha salva em: {output_file}")


def main():
    logger.info("=" * 90)
    logger.info("              AUTOMACAO DE PING - VERIFICADOR DE HOSTS")
    logger.info("=" * 90)
    
    # Arquivo hosts
    hosts_file = "hosts"
    
    logger.info(f"Lendo arquivo: {hosts_file}")
    entries = parse_hosts_file(hosts_file)
    
    # Remove duplicatas mantendo ordem
    seen = set()
    unique_entries = []
    for entry in entries:
        key = (entry[0], entry[1])  # ip, domain
        if key not in seen:
            seen.add(key)
            unique_entries.append(entry)
    
    total = len(unique_entries)
    logger.info(f"Total de entradas encontradas: {total}")
    
    # Mostra categorias encontradas
    categories = set(e[2] for e in unique_entries)
    logger.info(f"Categorias encontradas: {len(categories)}")
    for cat in sorted(categories):
        count = sum(1 for e in unique_entries if e[2] == cat)
        logger.info(f"  - {cat}: {count} hosts")
    
    logger.info("-" * 90)
    logger.info("Iniciando verificacao de conectividade (execucao paralela)...")
    logger.info("-" * 90)
    
    # Prepara argumentos para execução paralela
    args_list = [(i, total, ip, domain, category) for i, (ip, domain, category) in enumerate(unique_entries, 1)]
    
    # Executa pings em paralelo (max 50 threads para não sobrecarregar)
    results = []
    max_workers = min(50, total)  # Usa até 50 threads ou o total de hosts
    
    start_time = datetime.now()
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(ping_with_info, args): args for args in args_list}
        
        for future in as_completed(futures):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                logger.error(f"Erro ao processar: {e}")
    
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    
    # Gera nome do arquivo com timestamp
    output_file = f"resultado_ping_{timestamp}.xlsx"
    
    # Cria planilha
    logger.info("-" * 90)
    logger.info("Gerando planilha Excel...")
    logger.info("-" * 90)
    create_excel_report(results, output_file)
    
    # Resumo final
    online = sum(1 for r in results if r['success'])
    offline = len(results) - online
    
    logger.info("=" * 90)
    logger.info("                              RESUMO FINAL")
    logger.info("=" * 90)
    logger.info(f"  Total de hosts verificados: {len(results)}")
    logger.info(f"  [+] Online:  {online}")
    logger.info(f"  [-] Offline: {offline}")
    logger.info(f"  Tempo total: {duration:.2f} segundos")
    logger.info(f"  Arquivo de log: {log_filename}")
    logger.info(f"  Planilha: {output_file}")
    logger.info("=" * 90)


if __name__ == "__main__":
    main()
